from openpyxl import Workbook, load_workbook

pathURL = "Excel Files/JobURLs.XLSX"
pathJobsApplied = "Excel Files/JobsApplied.XLSX"


def append_job_information(data):

    wb = load_workbook(pathURL)
    ws = wb.active

    ws.append(list(data.values()))

    wb.save(pathURL)

def add_job_applied(data):

    wb = load_workbook(pathJobsApplied)
    ws = wb.active

    ws.append(list(data.values()))

    wb.save(pathJobsApplied)

def grab_and_delete_row(wb, ws):
    # Iterate over each row, print its value, then delete it.
    while(ws.max_row > 1):
        print("\n")
        
        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=2, column=j)
            if(j == 1):
                print("Company Name: ", cell_obj.value, end=" ")
            elif(j == 2):
                print("Job Title: ", cell_obj.value, end=" ")
            elif(j == 3):
                print("URL: ", cell_obj.value, end=" ")

        print("\nDeleting \n")
        ws.delete_rows(2)
    wb.save(pathURL)